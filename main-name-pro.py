import json
from seatable_api import Base
from openpyxl import load_workbook
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
import os
import shutil  # 用于复制文件
import glob

# 加载 .env 文件中的环境变量
load_dotenv()

# 全局变量存储当前选择的配置
current_config = None
current_seatable_config = None

def get_seatable_config_for_file(config_filename):
    """根据配置文件名获取对应的 SeaTable 配置"""
    # 移除文件扩展名并转换为环境变量格式
    base_name = os.path.splitext(config_filename)[0].upper().replace('-', '_')
    
    # 尝试获取特定配置文件的 API token
    specific_token_key = f"{base_name}_SEATABLE_API_TOKEN"
    api_token = os.getenv(specific_token_key)
    
    # 如果没有特定的 token，使用默认的
    if not api_token:
        api_token = os.getenv('DEFAULT_SEATABLE_API_TOKEN')
    
    return {
        'server_url': os.getenv('SEATABLE_SERVER_URL'),
        'api_token': api_token
    }

def load_json_configs():
    """自动加载目录下所有的 JSON 配置文件"""
    json_files = glob.glob('*.json')
    configs = []
    
    for json_file in sorted(json_files):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # 添加文件名信息到配置中
                config['config_filename'] = json_file
                config['config_display_name'] = os.path.splitext(json_file)[0]
                
                # 生成菜单显示名称：描述 + 文件名
                menu_description = config.get('menu_description', '')
                if menu_description:
                    config['menu_display_name'] = f"{menu_description} ({config['config_display_name']})"
                else:
                    config['menu_display_name'] = config['config_display_name']
                
                configs.append(config)
                print(f"已加载配置文件: {json_file}")
        except Exception as e:
            print(f"加载配置文件 {json_file} 失败: {e}")
    
    return configs

def get_column_index_by_name(sheet, column_name):
    """根据列名获取 Excel 中的列索引，假设第1行是标题，第2行是表头"""
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=2, column=col).value == column_name:
            return col
    raise ValueError(f"Column '{column_name}' not found in Excel sheet.")

def get_seatable_field_mapping(base, table_name):
    """获取 Seatable 表的所有元数据，并返回字段名称列表"""
    metadata = base.get_metadata()  # 获取所有表格的元数据
    print("Seatable Metadata:", json.dumps(metadata, indent=4, ensure_ascii=False))  # 打印调试元数据

    # 查找目标表的元数据
    table_metadata = None
    for table in metadata['tables']:
        if table['name'] == table_name:  # 根据表名匹配
            table_metadata = table
            break

    if table_metadata is None:
        raise ValueError(f"Table '{table_name}' not found in Seatable metadata.")

    # 返回字段名称列表
    return [col['name'] for col in table_metadata['columns']]

def sync_xlsx_for_table(table_config):
    # 获取系统当前日期并格式化为 YYYYMMDD 格式
    date_version = datetime.now().strftime('%Y%m%d')
    
    table_name = table_config['table_name']
    
    # 连接到 Seatable
    print(f"Connecting to Seatable: {current_seatable_config['server_url']}...")
    base = Base(current_seatable_config['api_token'], current_seatable_config['server_url'])
    base.auth()

    # 获取该表的字段名称列表
    seatable_field_names = get_seatable_field_mapping(base, table_name)
    print(f"Field Names for table '{table_name}':", seatable_field_names)

    # 获取 Seatable 数据
    print(f"Fetching data from Seatable table '{table_name}'...")
    rows = base.list_rows(table_name)
    seatable_df = pd.DataFrame(rows)
    print("Seatable DataFrame Columns:", seatable_df.columns)

    # 获取 relation_field 对应的字段名称
    relation_field_name = table_config['relation_field']
    relation_field_mappings = table_config.get('relation_field_mappings', {})
    if relation_field_name not in relation_field_mappings:
        raise KeyError(f"Relation field mapping for '{relation_field_name}' not found in relation_field_mappings.")
    
    # 获取 Excel 中的关联字段名
    excel_relation_field_name = relation_field_mappings[relation_field_name]

    # 检查 DataFrame 中的列是否包含该字段名称
    if relation_field_name not in seatable_df.columns:
        raise KeyError(f"Field '{relation_field_name}' not found in DataFrame columns.")

    # 转换关联字段为字符串
    seatable_df[relation_field_name] = seatable_df[relation_field_name].astype(str)

    # 读取 Excel 文件
    original_file_name = table_config['excel_file_name']
    excel_file_path = os.path.join(table_config['excel_directory'], f"{original_file_name}.xlsx")

    print(f"Loading Excel file '{original_file_name}.xlsx'...")
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook[table_config['sheet_name']]

    # 获取 Excel 中关联字段列的索引
    excel_relation_col = get_column_index_by_name(sheet, excel_relation_field_name)

    # 更新 Excel
    print(f"Starting to update Excel for table '{table_name}'...")
    updated_rows_count = 0

    # 遍历 Seatable 数据并找到匹配的 Excel 行
    for index, seatable_row in seatable_df.iterrows():  
        seatable_relation_value = str(seatable_row[relation_field_name]).strip()
    
        for excel_row in range(3, sheet.max_row + 1):  
            excel_relation_value = str(sheet.cell(row=excel_row, column=excel_relation_col).value).strip()
            print(f"Matching Seatable Value: '{seatable_relation_value}' with Excel Value: '{excel_relation_value}'")
            
            if seatable_relation_value == excel_relation_value:
                print(f"Match found! Updating Row {excel_row}")
                for seatable_field, excel_col_name in table_config['field_mappings'].items():
                    if seatable_field in seatable_df.columns:
                        excel_col = get_column_index_by_name(sheet, excel_col_name)
                        excel_value = seatable_row[seatable_field]
                        
                        if excel_value is not None:
                            # 输出调试信息
                            print(f"Updating Excel: Row {excel_row}, Column {excel_col}, Value {excel_value}")
                            
                            # 更新单元格值
                            sheet.cell(row=excel_row, column=excel_col).value = excel_value
                            updated_rows_count += 1

    # 保存 Excel 文件
    print(f"Saving changes to the original Excel file: {excel_file_path}...")
    workbook.save(filename=excel_file_path)

    # 创建带日期版本的副本文件
    file_name_with_date = f"{original_file_name}@{date_version}.xlsx"
    file_path_with_date = os.path.join(table_config['excel_directory'], file_name_with_date)
    
    print(f"Copying to '{file_name_with_date}'...")
    shutil.copy(excel_file_path, file_path_with_date)

    print(f"Completed! Total of {updated_rows_count} rows were updated for table '{table_name}'.")

def main_menu():
    """主菜单"""
    global current_config, current_seatable_config
    
    while True:
        configs = load_json_configs()
        
        if not configs:
            print("未找到任何 JSON 配置文件！")
            break
        
        print("\n=== SeaTable Excel 同步工具 ===")
        print("可用的配置文件:")
        for idx, config in enumerate(configs, start=1):
            print(f"{idx}. {config['menu_display_name']} (包含 {len(config['tables'])} 个表格)")
        print("0. 退出")
        
        try:
            choice = int(input("\n请选择配置文件: "))
            
            if choice == 0:
                print("退出程序")
                break
            elif 1 <= choice <= len(configs):
                current_config = configs[choice - 1]
                current_seatable_config = get_seatable_config_for_file(current_config['config_filename'])
                
                # 验证 SeaTable 配置
                if not current_seatable_config['api_token']:
                    print(f"警告: 未找到配置文件 {current_config['config_filename']} 对应的 API Token")
                    print("请检查 .env 文件中的配置")
                    continue
                
                print(f"\n已选择配置: {current_config['menu_display_name']}")
                print(f"SeaTable 服务器: {current_seatable_config['server_url']}")
                
                # 直接显示表格菜单
                table_menu()
            else:
                print("无效的选择，请重试")
        except ValueError:
            print("请输入有效的数字")

def table_menu():
    """显示表格菜单，让用户选择要处理的表格"""
    while True:
        print(f"\n配置文件: {current_config['menu_display_name']}")
        print("可用的表格:")
        tables = current_config['tables']
        
        # 显示所有可用的表格名称
        for idx, table_config in enumerate(tables, start=1):
            print(f"{idx}. {table_config['table_name']}")
        print(f"{len(tables) + 1}. 同步所有表格")
        print("0. 返回配置文件选择")

        try:
            choice = int(input("\n请输入表格编号: "))
            
            if choice == 0:
                break  # 返回主菜单
            elif 1 <= choice <= len(tables):
                table_config = tables[choice - 1]
                sync_xlsx_for_table(table_config)
                # 同步完成后继续显示表格菜单
            elif choice == len(tables) + 1:
                # 同步所有表格
                for table_config in tables:
                    print(f"\n开始同步表格: {table_config['table_name']}")
                    sync_xlsx_for_table(table_config)
                print("\n所有表格同步完成！")
                # 同步完成后继续显示表格菜单
            else:
                print("无效的选择，请重试")
        except ValueError:
            print("请输入有效的数字")

if __name__ == '__main__':
    print("正在加载配置...")
    main_menu()